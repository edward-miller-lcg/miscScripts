#!/usr/bin/env python3
"""
FHIR Resource Analyzer — QI Core 6.0.0

Recursively scans a directory for .ndjson files (one FHIR resource per line)
and produces a multi-sheet Excel workbook covering:

  • Summary              — per-resource-type overview
  • FHIR Element Analysis— QI Core element presence / code-system breakdown
  • Medication Codes     — per-code frequency table across medication resources
  • Category Codes       — category coding breakdown per resource type
  • Data Inventory       — exhaustive element-path frequency across all resources

Usage:
    python fhir_analyzer.py <root_dir> [-o output.xlsx]
                            [--facility NAME] [--period MMM_YYYY]
                            [--dd path/to/data_dictionary.xlsx]

Requirements:
    pip install openpyxl
"""

import json
import sys
import argparse
from pathlib import Path
from collections import defaultdict, Counter

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required.  Run:  pip install openpyxl")
    sys.exit(1)


# ── Code-system constants ─────────────────────────────────────────────────────

# Strict set: the three systems highlighted in QI Core medication / lab analytics
STANDARD_SYSTEMS = {
    "http://www.nlm.nih.gov/research/umls/rxnorm",   # RxNorm
    "http://snomed.info/sct",                          # SNOMED CT
    "http://loinc.org",                                # LOINC
}

# Broader set: any system the field considers "recognized / standardized"
# Matches the list used in the companion DQR script + hl7.org catch-all
RECOGNIZED_SYSTEMS = {
    "http://terminology.hl7.org/CodeSystem/v3-ActCode",
    "http://terminology.hl7.org/CodeSystem/v3-ActPriority",
    "http://terminology.hl7.org/CodeSystem/v3-ActUSPrivacyLaw",
    "http://terminology.hl7.org/CodeSystem/v2-0371",
    "http://hl7.org/fhir/R4/codesystem-address-type.html",
    "http://hl7.org/fhir/R4/valueset-address-use.html",
    "http://hl7.org/fhir/R4/codesystem-administrative-gender.html",
    "http://terminology.hl7.org/CodeSystem/v3-AdministrativeGender",
    "http://terminology.hl7.org/CodeSystem/admit-source",
    "urn:ietf:bcp:47",
    "http://terminology.hl7.org/CodeSystem/v2-0116",
    "http://hl7.org/fhir/R4/valueset-bundle-type.html",
    "urn:oid:2.16.840.1.113883.6.238",
    "http://terminology.hl7.org/CodeSystem/common-tags",
    "http://terminology.hl7.org/CodeSystem/condition-category",
    "http://terminology.hl7.org/CodeSystem/contactentity-type",
    "http://hl7.org/fhir/contact-point-system",
    "http://hl7.org/fhir/contact-point-use",
    "http://terminology.hl7.org/CodeSystem/coverage-class",
    "http://terminology.hl7.org/CodeSystem/coverage-copay-type",
    "http://www.ama-assn.org/go/cpt",
    "http://terminology.hl7.org/CodeSystem/data-absent-reason",
    "http://hl7.org/fhir/days-of-week",
    "http://hl7.org/fhir/device-nametype",
    "http://hl7.org/fhir/device-status",
    "http://terminology.hl7.org/CodeSystem/diagnosis-role",
    "http://hl7.org/fhir/diagnostic-report-status",
    "http://terminology.hl7.org/CodeSystem/diet",
    "http://terminology.hl7.org/CodeSystem/dose-rate-type",
    "http://hl7.org/fhir/encounter-location-status",
    "http://hl7.org/fhir/encounter-status",
    "http://hl7.org/fhir/event-status",
    "http://terminology.hl7.org/CodeSystem/list-example-use-codes",
    "http://hl7.org/fhir/R4/codesystem-device-status.html",
    "http://terminology.hl7.org/CodeSystem/device-status-reason",
    "http://hl7.org/fhir/fm-status",
    "https://www.cms.gov/Medicare/Coding/HCPCSReleaseCodeSets",
    "urn:oid:2.16.840.1.113883.6.259",
    "http://hl7.org/fhir/sid/icd-10-cm",
    "http://www.cms.gov/Medicare/Coding/ICD10",
    "http://hl7.org/fhir/sid/icd-9-cm",
    "http://hl7.org/fhir/R4/codesystem-identifier-use.html",
    "http://hl7.org/fhir/R4/codesystem-link-type.html",
    "http://hl7.org/fhir/ValueSet/list-mode",
    "http://hl7.org/fhir/ValueSet/list-status",
    "http://hl7.org/fhir/R4/codesystem-location-mode.html",
    "http://hl7.org/fhir/location-status",
    "http://terminology.hl7.org/CodeSystem/location-physical-type",
    "http://loinc.org",
    "http://terminology.hl7.org/CodeSystem/measure-improvement-notation",
    "http://terminology.hl7.org/CodeSystem/measure-population",
    "http://hl7.org/fhir/measure-report-status",
    "http://hl7.org/fhir/measure-report-type",
    "http://hl7.org/fhir/CodeSystem/medication-status",
    "http://terminology.hl7.org/CodeSystem/medicationrequest-category",
    "http://terminology.hl7.org/CodeSystem/medicationrequest-course-of-therapy",
    "http://hl7.org/fhir/R4/codesystem-medicationrequest-intent.html",
    "http://hl7.org/fhir/R4/codesystem-medicationrequest-status.html",
    "http://hl7.org/fhir/R4/codesystem-medicationrequest-status-reason.html",
    "http://hl7.org/fhir/name-use",
    "http://terminology.hl7.org/CodeSystem/observation-category",
    "http://terminology.hl7.org/CodeSystem/referencerange-meaning",
    "http://hl7.org/fhir/observation-status",
    "http://terminology.hl7.org/CodeSystem/organization-type",
    "urn:oid:1.2.36.1.2001.1001.101.104.16592",
    "http://terminology.hl7.org/CodeSystem/v2-0092",
    "http://terminology.hl7.org/CodeSystem/v2-0916",
    "http://hl7.org/fhir/request-intent",
    "http://hl7.org/fhir/request-priority",
    "http://hl7.org/fhir/request-status",
    "http://terminology.hl7.org/CodeSystem/v3-RoleCode",
    "http://www.nlm.nih.gov/research/umls/rxnorm",
    "http://terminology.hl7.org/CodeSystem/service-type",
    "http://snomed.info/sct",
    "https://nahdo.org/sopt",
    "http://terminology.hl7.org/CodeSystem/encounter-special-arrangements",
    "http://terminology.hl7.org/CodeSystem/v2-0493",
    "http://hl7.org/fhir/specimen-status",
    "http://terminology.hl7.org/CodeSystem/subscriber-relationship",
    "http://terminology.hl7.org/CodeSystem/v3-substanceAdminSubstitution",
    "http://terminology.hl7.org/CodeSystem/v3-TribalEntityUS",
    "http://unitsofmeasure.org",
    "http://hl7.org/fhir/udi-entry-type",
    "http://hl7.org/fhir/us/core/CodeSystem/us-core-category",
    "https://www.usps.com/",
    "http://terminology.hl7.org/CodeSystem/v2-0131",
    "http://terminology.hl7.org/CodeSystem/v2-0203",
    "http://terminology.hl7.org/CodeSystem/v2-0373",
    "http://terminology.hl7.org/CodeSystem/v3-EncounterSpecialCourtesy",
    "http://terminology.hl7.org/CodeSystem/v3-ObservationInterpretation",
    "http://terminology.hl7.org/CodeSystem/v3-MaritalStatus",
    "http://terminology.hl7.org/CodeSystem/v3-NullFlavor",
}


def is_recognized(system: str) -> bool:
    """True if system is in the recognized set OR contains 'hl7.org'."""
    return system in RECOGNIZED_SYSTEMS or "hl7.org" in system


def profile_short_name(url: str) -> str:
    """Return the last path segment of a profile URL for compact display."""
    return url.rstrip("/").split("/")[-1] if url else ""


# ── Field-presence helpers ───────────────────────────────────────────────────

def field_present(obj, path: str) -> bool:
    if obj is None:
        return False
    if not path:
        return obj not in (None, [], "")
    key, _, rest = path.partition(".")
    if isinstance(obj, list):
        return any(field_present(item, path) for item in obj)
    if not isinstance(obj, dict):
        return False
    val = obj.get(key)
    if val is None or val == [] or val == "":
        return False
    if not rest:
        return True
    return field_present(val, rest)


def poly_present(obj, *fields: str) -> bool:
    if not isinstance(obj, dict):
        return False
    return any(obj.get(f) not in (None, [], "") for f in fields)


def extension_present(resource: dict, url: str) -> bool:
    return any(
        isinstance(e, dict) and e.get("url") == url
        for e in resource.get("extension", [])
    )


def category_lab_present(resource: dict) -> bool:
    for cat in resource.get("category", []):
        if not isinstance(cat, dict):
            continue
        for coding in cat.get("coding", []):
            if isinstance(coding, dict) and coding.get("code") == "laboratory":
                return True
    return False


def component_value_present(resource: dict) -> bool:
    VALUE_X = (
        "valueQuantity", "valueCodeableConcept", "valueString", "valueBoolean",
        "valueInteger", "valueRange", "valueRatio", "valueSampledData",
        "valueTime", "valueDateTime", "valuePeriod",
    )
    comps = resource.get("component")
    if not comps:
        return False
    return any(poly_present(c, *VALUE_X) for c in comps if isinstance(c, dict))


def collected_x_present(resource: dict) -> bool:
    col = resource.get("collection")
    if not isinstance(col, dict):
        return False
    return poly_present(col, "collectedDateTime", "collectedPeriod")


# ── Code-system extraction helpers ───────────────────────────────────────────

def navigate(obj, path: str):
    if not path or obj is None:
        return obj
    key, _, rest = path.partition(".")
    if isinstance(obj, list):
        results = []
        for item in obj:
            r = navigate(item, path)
            if r is not None:
                results.extend(r if isinstance(r, list) else [r])
        return results if results else None
    if not isinstance(obj, dict):
        return None
    val = obj.get(key)
    if val is None:
        return None
    if not rest:
        return val
    return navigate(val, rest)


def systems_from_cc(val) -> list:
    if val is None:
        return []
    items = val if isinstance(val, list) else [val]
    out = []
    for item in items:
        if isinstance(item, dict):
            for coding in item.get("coding", []):
                if isinstance(coding, dict) and coding.get("system"):
                    out.append(coding["system"])
    return out


def cc_systems(resource: dict, path: str) -> list:
    return systems_from_cc(navigate(resource, path))


def poly_cc_systems(resource: dict, *cc_field_names: str) -> list:
    out = []
    for f in cc_field_names:
        val = resource.get(f)
        if isinstance(val, dict):
            out.extend(systems_from_cc(val))
    return out


def extract_race_systems(resource: dict) -> list:
    out = []
    for ext in resource.get("extension", []):
        if not isinstance(ext, dict):
            continue
        if ext.get("url") == "http://hl7.org/fhir/us/core/StructureDefinition/us-core-race":
            for sub in ext.get("extension", []):
                if not isinstance(sub, dict):
                    continue
                if sub.get("url") in ("ombCategory", "detailed"):
                    vc = sub.get("valueCoding", {})
                    if isinstance(vc, dict) and vc.get("system"):
                        out.append(vc["system"])
    return out


def extract_ethnicity_systems(resource: dict) -> list:
    out = []
    for ext in resource.get("extension", []):
        if not isinstance(ext, dict):
            continue
        if ext.get("url") == "http://hl7.org/fhir/us/core/StructureDefinition/us-core-ethnicity":
            for sub in ext.get("extension", []):
                if not isinstance(sub, dict):
                    continue
                if sub.get("url") in ("ombCategory", "detailed"):
                    vc = sub.get("valueCoding", {})
                    if isinstance(vc, dict) and vc.get("system"):
                        out.append(vc["system"])
    return out


# ── Check definitions ────────────────────────────────────────────────────────

def chk(resource_type, element, value_set, check_fn,
        extract_systems_fn=None, is_status=False, status_path=None,
        expected_systems=None):
    """
    expected_systems: set of code system URIs that are considered correct /
    expected for this specific element binding.  Used to compute
    "N Using Expected System" — a per-element correctness metric that is
    more meaningful than the generic RxNorm/SNOMED/LOINC count for elements
    whose binding calls for a different system (e.g. SOPT for Coverage.type,
    LOINC for Observation.code, HL7 v3 for category codes).
    """
    return dict(
        resource_type=resource_type, element=element, value_set=value_set,
        check_fn=check_fn, extract_systems_fn=extract_systems_fn,
        is_status=is_status, status_path=status_path,
        expected_systems=frozenset(expected_systems) if expected_systems else None,
    )


_RXNORM  = {"http://www.nlm.nih.gov/research/umls/rxnorm"}
_LOINC   = {"http://loinc.org"}
_SNOMED  = {"http://snomed.info/sct"}
_SOPT    = {"https://nahdo.org/sopt"}
_OMB     = {"urn:oid:2.16.840.1.113883.6.238"}
_OBS_CAT = {"http://terminology.hl7.org/CodeSystem/observation-category"}
_MR_CAT  = {"http://terminology.hl7.org/CodeSystem/medicationrequest-category"}
_ADMIT   = {"http://terminology.hl7.org/CodeSystem/admit-source"}
_DISCH   = {"http://terminology.hl7.org/CodeSystem/discharge-disposition"}

CHECKS = [
    # ── Coverage ─────────────────────────────────────────────────────────────
    chk("Coverage", "Coverage.period", "",
        lambda r: field_present(r, "period")),
    chk("Coverage", "Coverage.status",
        "http://hl7.org/fhir/ValueSet/fm-status|4.0.1",
        lambda r: field_present(r, "status"),
        is_status=True, status_path="status"),
    chk("Coverage", "Coverage.type",
        "http://cts.nlm.nih.gov/fhir/ValueSet/2.16.840.1.114222.4.11.3591",
        lambda r: field_present(r, "type"),
        extract_systems_fn=lambda r: cc_systems(r, "type"),
        expected_systems=_SOPT),

    # ── Encounter ─────────────────────────────────────────────────────────────
    chk("Encounter", "Encounter.hospitalization.admitSource",
        "https://hl7.org/fhir/R4/valueset-encounter-admit-source.html",
        lambda r: field_present(r, "hospitalization.admitSource"),
        extract_systems_fn=lambda r: cc_systems(r, "hospitalization.admitSource"),
        expected_systems=_ADMIT),
    chk("Encounter", "Encounter.hospitalization.dischargeDisposition",
        "https://terminology.hl7.org/6.1.0/ValueSet-clinical-discharge-disposition.html",
        lambda r: field_present(r, "hospitalization.dischargeDisposition"),
        extract_systems_fn=lambda r: cc_systems(r, "hospitalization.dischargeDisposition"),
        expected_systems=_DISCH),
    chk("Encounter", "Encounter.period", "",
        lambda r: field_present(r, "period")),
    chk("Encounter", "Encounter.period.end", "",
        lambda r: field_present(r, "period.end")),
    chk("Encounter", "Encounter.period.start", "",
        lambda r: field_present(r, "period.start")),

    # ── Medication ────────────────────────────────────────────────────────────
    chk("Medication", "Medication.code",
        "http://cts.nlm.nih.gov/fhir/ValueSet/2.16.840.1.113762.1.4.1010.4",
        lambda r: field_present(r, "code"),
        extract_systems_fn=lambda r: cc_systems(r, "code"),
        expected_systems=_RXNORM),
    chk("Medication", "Medication.status",
        "http://hl7.org/fhir/R4/valueset-medication-status.html",
        lambda r: field_present(r, "status"),
        is_status=True, status_path="status"),

    # ── MedicationAdministration ──────────────────────────────────────────────
    chk("MedicationAdministration", "MedicationAdministration.dosage.route",
        "http://hl7.org/fhir/ValueSet/route-codes",
        lambda r: field_present(r, "dosage.route"),
        extract_systems_fn=lambda r: cc_systems(r, "dosage.route"),
        expected_systems=_SNOMED),
    chk("MedicationAdministration", "MedicationAdministration.effective[x]", "",
        lambda r: poly_present(r, "effectiveDateTime", "effectivePeriod")),
    chk("MedicationAdministration", "MedicationAdministration.effectivePeriod.end", "",
        lambda r: field_present(r, "effectivePeriod.end")),
    chk("MedicationAdministration", "MedicationAdministration.effectivePeriod.start", "",
        lambda r: field_present(r, "effectivePeriod.start")),
    chk("MedicationAdministration", "MedicationAdministration.medication[x]",
        "https://vsac.nlm.nih.gov/valueset/2.16.840.1.113762.1.4.1010.4/expansion",
        lambda r: poly_present(r, "medicationCodeableConcept", "medicationReference"),
        extract_systems_fn=lambda r: poly_cc_systems(r, "medicationCodeableConcept"),
        expected_systems=_RXNORM),

    # ── MedicationRequest ─────────────────────────────────────────────────────
    chk("MedicationRequest", "MedicationRequest.authoredOn", "",
        lambda r: field_present(r, "authoredOn")),
    chk("MedicationRequest", "MedicationRequest.category",
        "http://hl7.org/fhir/ValueSet/medicationrequest-category",
        lambda r: field_present(r, "category"),
        extract_systems_fn=lambda r: cc_systems(r, "category"),
        expected_systems=_MR_CAT),
    chk("MedicationRequest", "MedicationRequest.dosageInstruction.route",
        "https://hl7.org/fhir/valueset-route-codes.html",
        lambda r: field_present(r, "dosageInstruction.route"),
        extract_systems_fn=lambda r: cc_systems(r, "dosageInstruction.route"),
        expected_systems=_SNOMED),
    chk("MedicationRequest", "MedicationRequest.dosageInstruction.Timing", "",
        lambda r: field_present(r, "dosageInstruction.timing")),
    chk("MedicationRequest", "MedicationRequest.dosageInstruction.Timing.boundsPeriod.end", "",
        lambda r: field_present(r, "dosageInstruction.timing.repeat.boundsPeriod.end")),
    chk("MedicationRequest", "MedicationRequest.dosageInstruction.Timing.boundsPeriod.start", "",
        lambda r: field_present(r, "dosageInstruction.timing.repeat.boundsPeriod.start")),
    chk("MedicationRequest", "MedicationRequest.medication[x]",
        "http://cts.nlm.nih.gov/fhir/ValueSet/2.16.840.1.113762.1.4.1010.4",
        lambda r: poly_present(r, "medicationCodeableConcept", "medicationReference"),
        extract_systems_fn=lambda r: poly_cc_systems(r, "medicationCodeableConcept"),
        expected_systems=_RXNORM),
    chk("MedicationRequest", "MedicationRequest.requester", "",
        lambda r: field_present(r, "requester")),

    # ── Observation ───────────────────────────────────────────────────────────
    chk("Observation", "Observation.category",
        "http://hl7.org/fhir/valueset-observation-category.html",
        lambda r: field_present(r, "category"),
        extract_systems_fn=lambda r: cc_systems(r, "category"),
        expected_systems=_OBS_CAT),
    chk("Observation", "Observation.category:Laboratory",
        "http://hl7.org/fhir/us/core/ValueSet-us-core-clinical-result-observation-category.html",
        category_lab_present,
        extract_systems_fn=lambda r: cc_systems(r, "category"),
        expected_systems=_OBS_CAT),
    chk("Observation", "Observation.code",
        "http://hl7.org/fhir/us/core/ValueSet/us-core-laboratory-test-codes",
        lambda r: field_present(r, "code"),
        extract_systems_fn=lambda r: cc_systems(r, "code"),
        expected_systems=_LOINC),
    chk("Observation", "Observation.component", "",
        lambda r: field_present(r, "component")),
    chk("Observation", "Observation.component.code",
        "http://hl7.org/fhir/ValueSet/observation-codes",
        lambda r: field_present(r, "component.code"),
        extract_systems_fn=lambda r: cc_systems(r, "component.code"),
        expected_systems=_LOINC),
    chk("Observation", "Observation.component.value[x]", "",
        component_value_present),
    chk("Observation", "Observation.effective[x]", "",
        lambda r: poly_present(r, "effectiveDateTime", "effectivePeriod",
                               "effectiveInstant", "effectiveTiming")),
    chk("Observation", "Observation.status",
        "http://hl7.org/fhir/us/qicore/ValueSet-qicore-non-negative-observation-status.html",
        lambda r: field_present(r, "status"),
        is_status=True, status_path="status"),
    chk("Observation", "Observation.subject", "",
        lambda r: field_present(r, "subject")),
    chk("Observation", "Observation.value[x]", "",
        lambda r: poly_present(
            r, "valueQuantity", "valueCodeableConcept", "valueString",
            "valueBoolean", "valueInteger", "valueRange", "valueRatio",
            "valueSampledData", "valueTime", "valueDateTime", "valuePeriod",
        )),

    # ── Patient ───────────────────────────────────────────────────────────────
    chk("Patient", "Patient.birthDate", "",
        lambda r: field_present(r, "birthDate")),
    chk("Patient", "Patient.extension (race)",
        "http://hl7.org/fhir/us/core/StructureDefinition/us-core-race",
        lambda r: extension_present(r, "http://hl7.org/fhir/us/core/StructureDefinition/us-core-race"),
        extract_systems_fn=extract_race_systems,
        expected_systems=_OMB),
    chk("Patient", "Patient.extension (sex at birth)",
        "http://hl7.org/fhir/us/core/StructureDefinition/us-core-birthsex",
        lambda r: extension_present(r, "http://hl7.org/fhir/us/core/StructureDefinition/us-core-birthsex")),
    chk("Patient", "Patient.extension:ethnicity",
        "https://hl7.org/fhir/us/core/STU6.1/ValueSet-omb-ethnicity-category.html",
        lambda r: extension_present(r, "http://hl7.org/fhir/us/core/StructureDefinition/us-core-ethnicity"),
        extract_systems_fn=extract_ethnicity_systems,
        expected_systems=_OMB),
    chk("Patient", "Patient.identifier", "",
        lambda r: field_present(r, "identifier")),
    chk("Patient", "Patient.name", "",
        lambda r: field_present(r, "name")),
    chk("Patient", "Patient.name.family", "",
        lambda r: field_present(r, "name.family")),
    chk("Patient", "Patient.name.given", "",
        lambda r: field_present(r, "name.given")),

    # ── Specimen ──────────────────────────────────────────────────────────────
    chk("Specimen", "Specimen.collection", "",
        lambda r: field_present(r, "collection")),
    chk("Specimen", "Specimen.collection.bodySite",
        "http://hl7.org/fhir/valueset-body-site.html",
        lambda r: field_present(r, "collection.bodySite"),
        extract_systems_fn=lambda r: cc_systems(r, "collection.bodySite"),
        expected_systems=_SNOMED),
    chk("Specimen", "Specimen.collection.collected[x]", "",
        collected_x_present),
    chk("Specimen", "Specimen.type",
        "https://vsac.nlm.nih.gov/valueset/2.16.840.1.113762.1.4.1099.54/expansion",
        lambda r: field_present(r, "type"),
        extract_systems_fn=lambda r: cc_systems(r, "type"),
        expected_systems=_SNOMED),
]


# ── File loading ─────────────────────────────────────────────────────────────

def find_ndjson_files(root: Path) -> list:
    return sorted(root.rglob("*.ndjson"))


def load_resources(files: list) -> tuple:
    resources = defaultdict(list)
    total_lines = parse_errors = 0
    for path in files:
        try:
            with open(path, "r", encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if not line:
                        continue
                    total_lines += 1
                    try:
                        resource = json.loads(line)
                        rtype = resource.get("resourceType")
                        if rtype:
                            resources[rtype].append(resource)
                    except json.JSONDecodeError:
                        parse_errors += 1
        except OSError as exc:
            print(f"  WARNING: cannot read {path}: {exc}")
    return dict(resources), total_lines, parse_errors


# ── Profile helpers ───────────────────────────────────────────────────────────

def collect_profile_counts(resources: dict) -> dict:
    """meta.profile counts per resource type. Stores (short_name, full_url) pairs."""
    out = {}
    for rtype, rlist in resources.items():
        counts = defaultdict(int)
        for r in rlist:
            meta = r.get("meta")
            if isinstance(meta, dict):
                for p in meta.get("profile", []):
                    if isinstance(p, str):
                        counts[p] += 1
        out[rtype] = dict(counts)
    return out


def profile_strings(profile_counts: dict) -> tuple:
    """
    Returns two pipe-delimited strings using SHORT names:
      ("ShortName1 | ShortName2", "count1 | count2")
    Sorted by count descending.
    """
    if not profile_counts:
        return "", ""
    pairs = sorted(profile_counts.items(), key=lambda x: -x[1])
    return (
        " | ".join(profile_short_name(p) for p, _ in pairs),
        " | ".join(str(c) for _, c in pairs),
    )


# ── Core analysis ─────────────────────────────────────────────────────────────

def analyze(resources: dict) -> list:
    results = []
    for check in CHECKS:
        rtype = check["resource_type"]
        rlist = resources.get(rtype, [])
        total = len(rlist)

        if total == 0:
            results.append({
                **check,
                "total": 0, "missing": 0, "present": "N/A",
                "system_counts": {}, "status_counts": {},
                "n_standard": 0, "n_recognized": 0, "n_expected": 0,
            })
            continue

        missing      = 0
        sys_counts   = defaultdict(int)
        sta_counts   = defaultdict(int)
        n_standard   = 0
        n_recognized = 0
        n_expected   = 0
        exp_sys      = check.get("expected_systems")  # frozenset or None

        for r in rlist:
            if not check["check_fn"](r):
                missing += 1

            if check["extract_systems_fn"] and check["value_set"]:
                systems = check["extract_systems_fn"](r)
                for s in systems:
                    sys_counts[s] += 1
                if any(s in STANDARD_SYSTEMS for s in systems):
                    n_standard += 1
                if any(is_recognized(s) for s in systems):
                    n_recognized += 1
                if exp_sys and any(s in exp_sys for s in systems):
                    n_expected += 1

            if check["is_status"] and check["status_path"]:
                val = navigate(r, check["status_path"])
                if isinstance(val, str) and val:
                    sta_counts[val] += 1

        results.append({
            **check,
            "total":         total,
            "missing":       missing,
            "present":       "Yes" if missing == 0 else "No",
            "system_counts": dict(sys_counts),
            "status_counts": dict(sta_counts),
            "n_standard":    n_standard,
            "n_recognized":  n_recognized,
            "n_expected":    n_expected,
        })

    return results


def expand_result(result: dict, prof_used: str, prof_counts: str) -> list:
    total   = result["total"]
    has_vs  = bool(result["value_set"])
    is_stat = result["is_status"]
    has_ext = result["extract_systems_fn"] is not None

    base = {
        "resource_type":  result["resource_type"],
        "element":        result["element"],
        "value_set":      result["value_set"],
        "total":          total,
        "missing":        result["missing"],
        "present":        result["present"],
        "profiles_used":  prof_used,
        "profiles_count": prof_counts,
    }

    has_exp = bool(result.get("expected_systems"))

    if has_vs and not is_stat and has_ext:
        n_std = result["n_standard"]
        n_rec = result["n_recognized"]
        n_exp = result["n_expected"] if has_exp else None
        p_std = round(n_std / total, 4) if total else 0.0
        std = {
            "n_standard":   n_std,
            "n_recognized": n_rec,
            "n_expected":   n_exp,
            "proportion":   p_std,
            "opposite":     round(1.0 - p_std, 4),
        }
    elif has_vs and is_stat:
        std = {"n_standard": "N/A", "n_recognized": "N/A",
               "n_expected": "N/A", "proportion": "N/A", "opposite": "N/A"}
    else:
        std = {"n_standard": None, "n_recognized": None,
               "n_expected": None, "proportion": None, "opposite": None}

    rows = []

    if is_stat and result["status_counts"]:
        for sv, sc in sorted(result["status_counts"].items()):
            rows.append({**base, **std,
                         "code_system": None, "cs_count": None,
                         "status_value": sv,   "status_count": sc})
    elif has_vs and has_ext and result["system_counts"]:
        for sys, cnt in sorted(result["system_counts"].items(), key=lambda x: -x[1]):
            rows.append({**base, **std,
                         "code_system": sys,  "cs_count": cnt,
                         "status_value": None, "status_count": None})
    else:
        rows.append({**base, **std,
                     "code_system": None, "cs_count": None,
                     "status_value": None, "status_count": None})

    return rows


# ── Category codes ────────────────────────────────────────────────────────────

# Resource types that have additional/alternate "category-like" fields
# beyond the standard `.category` (CodeableConcept array).
_CATEGORY_FIELDS = {
    "Encounter":                ["category", "type", "class"],
    "Medication":               ["category", "code"],
    "MedicationRequest":        ["category", "medicationCodeableConcept"],
    "MedicationAdministration": ["category", "medicationCodeableConcept"],
}


def _iter_codings(value):
    """
    Yield (system, code, display) tuples from a value that may be a
    Coding, a CodeableConcept, or a list of either (e.g. Encounter.class
    is a single Coding, Encounter.type / medicationCodeableConcept are
    CodeableConcept).  For CodeableConcept, falls back to `.text` when an
    individual coding has no `display`.
    """
    if isinstance(value, list):
        for item in value:
            yield from _iter_codings(item)
    elif isinstance(value, dict):
        if "coding" in value:
            cc_text = (value.get("text") or "").strip()
            for coding in value.get("coding", []):
                if isinstance(coding, dict):
                    system  = (coding.get("system")  or "").strip()
                    code    = (coding.get("code")    or "").strip()
                    display = (coding.get("display") or "").strip() or cc_text
                    if system or code:
                        yield (system, code, display)
        elif "system" in value or "code" in value:
            system  = (value.get("system")  or "").strip()
            code    = (value.get("code")    or "").strip()
            display = (value.get("display") or "").strip()
            if system or code:
                yield (system, code, display)


def collect_category_codes(resources: dict) -> dict:
    """
    For each resource type, collect (field, system, code) tuples from
    category-like fields (category.coding[], plus per-type extras such
    as Encounter.type/class and the medication code fields on
    Medication / MedicationRequest / MedicationAdministration).
    Returns {rtype: {(field, system, code): {"count": n, "display": str}}}.
    """
    out = {}
    for rtype, rlist in resources.items():
        counts: dict = {}
        fields = _CATEGORY_FIELDS.get(rtype, ["category"])
        for r in rlist:
            for field in fields:
                for system, code, display in _iter_codings(r.get(field)):
                    key = (field, system, code)
                    entry = counts.setdefault(key, {"count": 0, "display": ""})
                    entry["count"] += 1
                    if display and not entry["display"]:
                        entry["display"] = display
        if counts:
            out[rtype] = counts
    return out


# ── Data inventory (exhaustive element flattening) ────────────────────────────

def _unpack(data, path: str, results: dict) -> None:
    """
    Recursively flatten a FHIR resource into {dot.separated.path: count}.
    Arrays are traversed without adding an index segment (matches DQR script
    behaviour — paths reflect structure, not position).
    """
    if isinstance(data, dict):
        for key, value in data.items():
            new_path = f"{path}.{key}" if path else key
            _unpack(value, new_path, results)
    elif isinstance(data, list):
        for item in data:
            _unpack(item, path, results)
    else:
        if path:
            results[path] = results.get(path, 0) + 1


def build_data_inventory(resources: dict) -> list:
    """
    Flatten every resource of every type and count element-path occurrences.
    Returns list of {resource_type, element_path, count} sorted by type then
    count descending.
    """
    inventory = {}   # {(rtype, path): count}

    for rtype, rlist in resources.items():
        for resource in rlist:
            flat = {}
            _unpack(resource, "", flat)
            for path, _ in flat.items():
                key = (rtype, path)
                inventory[key] = inventory.get(key, 0) + 1

    rows = [
        {"resource_type": rt, "element_path": path, "count": cnt}
        for (rt, path), cnt in inventory.items()
    ]
    rows.sort(key=lambda r: (r["resource_type"], -r["count"]))
    return rows


# ── Medication code frequency ─────────────────────────────────────────────────

MED_RESOURCE_TYPES = [
    "Medication",
    "MedicationRequest",
    "MedicationAdministration",
    "MedicationDispense",
]

_MED_CC_FIELDS = {
    "Medication":               ["code"],
    "MedicationRequest":        ["medicationCodeableConcept"],
    "MedicationAdministration": ["medicationCodeableConcept"],
    "MedicationDispense":       ["medicationCodeableConcept"],
}


def _build_medication_lookup(resources: dict) -> dict:
    """
    Build a {medication_id: [(system, code, display), ...]} map from all
    Medication resources.  Used to resolve medicationReference pointers on
    MedicationRequest / MedicationAdministration so that RxNorm (and other)
    codes stored on the referenced Medication are not missed.
    """
    lookup: dict = {}
    for med in resources.get("Medication", []):
        mid = (med.get("id", "") or "").rstrip("/").split("/")[-1]
        if not mid:
            continue
        codings = []
        cc = med.get("code")
        if isinstance(cc, dict):
            cc_text = (cc.get("text") or "").strip()
            for coding in cc.get("coding", []):
                if not isinstance(coding, dict):
                    continue
                system  = (coding.get("system")  or "").strip()
                code    = (coding.get("code")    or "").strip()
                display = (coding.get("display") or "").strip() or cc_text
                if system or code:
                    codings.append((system, code, display))
        lookup[mid] = codings
    return lookup


def _resolve_med_ref(resource: dict, med_lookup: dict) -> list:
    """
    If the resource carries a medicationReference, resolve it against
    med_lookup and return its (system, code, display) tuples.
    Returns [] when the reference is absent, unresolvable, or already
    covered by an inline medicationCodeableConcept.
    """
    med_ref = resource.get("medicationReference")
    if not isinstance(med_ref, dict):
        return []
    ref_str = med_ref.get("reference", "")
    # Extract the ID — handles "Medication/abc", full URLs, and bare IDs
    parts = ref_str.rstrip("/").split("/")
    mid = parts[-1] if parts else ""
    return med_lookup.get(mid, [])


def _extract_med_codings(resource: dict, rtype: str,
                          med_lookup: dict | None = None) -> list:
    """
    Return (system, code, display) tuples for a medication resource.

    Checks BOTH the inline CodeableConcept field AND any medicationReference
    so that RxNorm codes stored on a separate Medication resource are included.
    Deduplicates on (system, code) to avoid double-counting when both paths
    carry the same coding.
    """
    seen   = set()
    results = []

    def _add_from_cc(cc):
        if not isinstance(cc, dict):
            return
        cc_text = (cc.get("text") or "").strip()
        for coding in cc.get("coding", []):
            if not isinstance(coding, dict):
                continue
            system  = (coding.get("system")  or "").strip()
            code    = (coding.get("code")    or "").strip()
            display = (coding.get("display") or "").strip() or cc_text
            if (system or code) and (system, code) not in seen:
                seen.add((system, code))
                results.append((system, code, display))

    # 1 — inline CodeableConcept (e.g. medicationCodeableConcept / code)
    for field in _MED_CC_FIELDS.get(rtype, []):
        _add_from_cc(resource.get(field))

    # 2 — referenced Medication resource (carries the RxNorm code when the
    #     request uses medicationReference instead of an inline CC)
    if med_lookup is not None:
        for system, code, display in _resolve_med_ref(resource, med_lookup):
            if (system, code) not in seen:
                seen.add((system, code))
                results.append((system, code, display))

    return results


def build_med_code_counts(resources: dict) -> tuple:
    present_types = [t for t in MED_RESOURCE_TYPES if resources.get(t)]
    med_lookup    = _build_medication_lookup(resources)
    counts: dict  = defaultdict(lambda: defaultdict(int))

    for rtype in present_types:
        for resource in resources[rtype]:
            for key in _extract_med_codings(resource, rtype, med_lookup):
                counts[key][rtype] += 1

    rows = [
        {"system": s, "code": c, "display": d,
         "by_type": dict(tc), "total": sum(tc.values())}
        for (s, c, d), tc in counts.items()
    ]
    rows.sort(key=lambda r: (-r["total"], r["system"], r["code"], r["display"]))
    return rows, present_types


# ── Colour / style constants ──────────────────────────────────────────────────

_C = {
    "hdr_bg":      "1F4E79",
    "hdr_fg":      "FFFFFF",
    "stripe_a":    "D6E4F0",
    "stripe_b":    "FFFFFF",
    "yes_bg":      "C6EFCE",  "yes_fg":  "276221",
    "no_bg":       "FFC7CE",  "no_fg":   "9C0006",
    "na_bg":       "FFEB9C",  "na_fg":   "7D5A00",
    "std_bg":      "E2EFDA",
    "rec_bg":      "EBF3FB",  # softer blue for recognized-system highlight
    "exp_bg":      "FFF2CC",  # gold/amber for per-element expected-system column
    "summary_hdr": "2E75B6",
    "meta_bg":     "F2F2F2",
}

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
PCTFMT = "0.00%"
NUMFMT = "#,##0"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=None, size=11) -> Font:
    kw = {"bold": bold, "size": size}
    if color:
        kw["color"] = color
    return Font(**kw)


# ── Sheet: FHIR Element Analysis ──────────────────────────────────────────────

DETAIL_HEADERS = [
    "FHIR Resource\n(QI Core 6.0.0)",              # A
    "FHIR Data Element",                             # B
    "FHIR Value Set",                                # C
    "Total\nResources",                              # D
    "Count\nMissing",                                # E
    "Present",                                       # F
    "N Using\nRxNorm /\nSNOMED / LOINC",            # G  ← strict 3
    "N Using Any\nRecognized\nSystem",               # H  ← broader set
    "N Using\nExpected\nSystem",                     # I  ← per-element correct system
    "Proportion\nUsing RxNorm /\nSNOMED / LOINC",   # J
    "Opposite\nProportion",                          # K
    "Code System\nUsed",                             # L
    "Use\nCount",                                    # M
    "Status\nUsed",                                  # N
    "Status\nUse Count",                             # O
    "Profiles Used\n(short name)",                   # P
    "Profiles\nUse Count",                           # Q
]

DETAIL_COL_WIDTHS = [28, 50, 62, 12, 12, 10, 14, 14, 14, 14, 12, 52, 10, 18, 14, 38, 14]


def write_detail_sheet(ws, all_rows: list) -> None:
    ws.row_dimensions[1].height = 52

    for ci, h in enumerate(DETAIL_HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = _font(bold=True, color=_C["hdr_fg"])
        cell.fill      = _fill(_C["hdr_bg"])
        cell.alignment = CENTER

    prev_rtype   = None
    stripe_color = _C["stripe_a"]

    for ri, row in enumerate(all_rows, 2):
        ws.row_dimensions[ri].height = 16

        if row["resource_type"] != prev_rtype:
            stripe_color = _C["stripe_a"] if stripe_color == _C["stripe_b"] else _C["stripe_b"]
            prev_rtype = row["resource_type"]

        stripe = _fill(stripe_color)

        def put(col, val, align=LEFT, fmt=None):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.fill      = stripe
            cell.alignment = align
            cell.font      = _font()
            if fmt:
                cell.number_format = fmt
            return cell

        put(1,  row["resource_type"])
        put(2,  row["element"])
        put(3,  row["value_set"])
        put(4,  row["total"],   CENTER, NUMFMT)
        put(5,  row["missing"], CENTER, NUMFMT)

        # Present
        pval = row["present"]
        pc = ws.cell(row=ri, column=6, value=pval)
        pc.alignment = CENTER
        if pval == "Yes":
            pc.fill = _fill(_C["yes_bg"]); pc.font = _font(bold=True, color=_C["yes_fg"])
        elif pval == "No":
            pc.fill = _fill(_C["no_bg"]);  pc.font = _font(bold=True, color=_C["no_fg"])
        else:
            pc.fill = _fill(_C["na_bg"]);  pc.font = _font(bold=True, color=_C["na_fg"])

        # System-count columns G–K
        # exp_bg is gold/amber to distinguish "element-specific correctness"
        # from the generic standard (green) and recognized (blue) columns
        for col, key, fmt, hi_color in [
            (7,  "n_standard",   NUMFMT, _C["std_bg"]),
            (8,  "n_recognized", NUMFMT, _C["rec_bg"]),
            (9,  "n_expected",   NUMFMT, _C["exp_bg"]),
            (10, "proportion",   PCTFMT, _C["std_bg"]),
            (11, "opposite",     PCTFMT, _C["rec_bg"]),
        ]:
            val = row.get(key)
            cell = ws.cell(row=ri, column=col, value=val)
            cell.alignment = CENTER
            if val == "N/A":
                cell.fill = _fill(_C["na_bg"])
                cell.font = _font(color=_C["na_fg"])
            elif val is not None:
                cell.fill = _fill(hi_color)
                cell.font = _font()
                if fmt:
                    cell.number_format = fmt
            else:
                cell.fill = stripe
                cell.font = _font()

        put(12, row.get("code_system"),    LEFT)
        put(13, row.get("cs_count"),       CENTER, NUMFMT)
        put(14, row.get("status_value"),   LEFT)
        put(15, row.get("status_count"),   CENTER, NUMFMT)
        put(16, row.get("profiles_used"),  LEFT)
        put(17, row.get("profiles_count"), LEFT)

    for ci, w in enumerate(DETAIL_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


# ── Sheet: Summary ────────────────────────────────────────────────────────────

def write_summary_sheet(ws, results: list, all_resource_counts: dict,
                        facility: str, period: str) -> None:
    # Metadata rows at top
    meta_rows = []
    if facility:
        meta_rows.append(("Facility", facility))
    if period:
        meta_rows.append(("Reporting Period", period))
    meta_rows.append(("Generated", __import__("datetime").date.today().isoformat()))

    for mi, (label, value) in enumerate(meta_rows, 1):
        ws.row_dimensions[mi].height = 18
        lc = ws.cell(row=mi, column=1, value=label)
        vc = ws.cell(row=mi, column=2, value=value)
        for cell in (lc, vc):
            cell.fill = _fill(_C["meta_bg"])
            cell.font = _font(bold=(cell == lc))
            cell.alignment = LEFT

    hdr_row = len(meta_rows) + 2   # blank separator

    SUMM_HDR = [
        "Resource Type", "Total Resources", "Elements Checked",
        "Fully Present (Yes)", "Has Gaps (No)", "No Data (N/A)",
        "Distinct Code Systems", "Distinct Profiles",
    ]
    ws.row_dimensions[hdr_row].height = 28
    for ci, h in enumerate(SUMM_HDR, 1):
        cell = ws.cell(row=hdr_row, column=ci, value=h)
        cell.font      = _font(bold=True, color=_C["hdr_fg"])
        cell.fill      = _fill(_C["summary_hdr"])
        cell.alignment = CENTER

    by_type = defaultdict(list)
    for r in results:
        by_type[r["resource_type"]].append(r)

    for ri, rtype in enumerate(sorted(by_type), hdr_row + 1):
        rrows    = by_type[rtype]
        n_yes    = sum(1 for r in rrows if r["present"] == "Yes")
        n_no     = sum(1 for r in rrows if r["present"] == "No")
        n_na     = sum(1 for r in rrows if r["present"] == "N/A")
        all_sys  = set()
        for r in rrows:
            all_sys.update(r.get("system_counts", {}).keys())

        stripe = _fill(_C["stripe_a"]) if ri % 2 == 0 else _fill(_C["stripe_b"])
        vals = [
            rtype,
            all_resource_counts.get(rtype, 0),
            len(rrows), n_yes, n_no, n_na,
            len(all_sys),
            len(all_resource_counts.get("__profiles__" + rtype, {})),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill      = stripe
            cell.alignment = CENTER if ci > 1 else LEFT
            cell.font      = _font()

    for ci, w in enumerate([28, 16, 18, 18, 14, 14, 22, 18], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{hdr_row + 1}"


# ── Sheet: Medication Codes ───────────────────────────────────────────────────

def write_medication_sheet(ws, rows: list, present_types: list) -> None:
    headers = ["Code System", "Code", "Display"]
    for rtype in present_types:
        headers.append(f"Count in\n{rtype}")
    headers.append("Total\nCount")

    ws.row_dimensions[1].height = 40
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = _font(bold=True, color=_C["hdr_fg"])
        cell.fill      = _fill(_C["hdr_bg"])
        cell.alignment = CENTER

    if not rows:
        ws.cell(row=2, column=1, value="No medication coding data found.")
        return

    for ri, row in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 16
        if row["system"] in STANDARD_SYSTEMS:
            stripe = _fill(_C["std_bg"])
        elif ri % 2 == 0:
            stripe = _fill(_C["stripe_a"])
        else:
            stripe = _fill(_C["stripe_b"])

        def put(col, val, align=LEFT, fmt=None, bold=False):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.fill      = stripe
            cell.alignment = align
            cell.font      = _font(bold=bold)
            if fmt:
                cell.number_format = fmt

        put(1, row["system"])
        put(2, row["code"])
        put(3, row["display"])
        for ti, rtype in enumerate(present_types, 4):
            put(ti, row["by_type"].get(rtype, 0), CENTER, NUMFMT)
        total_col = 4 + len(present_types)
        tc = ws.cell(row=ri, column=total_col, value=row["total"])
        tc.fill = stripe; tc.alignment = CENTER
        tc.font = _font(bold=True); tc.number_format = NUMFMT

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 52
    for ci in range(4, 4 + len(present_types) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.freeze_panes = "A2"

    # Totals row
    totals_row = 1 + len(rows) + 2
    ws.row_dimensions[totals_row].height = 18
    ws.cell(row=totals_row, column=1, value="TOTALS").font = _font(bold=True, color=_C["hdr_fg"])
    ws.cell(row=totals_row, column=1).fill = _fill(_C["summary_hdr"])
    ws.cell(row=totals_row, column=1).alignment = CENTER
    for ti, rtype in enumerate(present_types, 4):
        ct = sum(r["by_type"].get(rtype, 0) for r in rows)
        cell = ws.cell(row=totals_row, column=ti, value=ct)
        cell.font = _font(bold=True, color=_C["hdr_fg"])
        cell.fill = _fill(_C["summary_hdr"])
        cell.alignment = CENTER; cell.number_format = NUMFMT
    gt_col = 4 + len(present_types)
    gt = ws.cell(row=totals_row, column=gt_col, value=sum(r["total"] for r in rows))
    gt.font = _font(bold=True, color=_C["hdr_fg"])
    gt.fill = _fill(_C["summary_hdr"])
    gt.alignment = CENTER; gt.number_format = NUMFMT


# ── Sheet: Category Codes ─────────────────────────────────────────────────────

def write_category_codes_sheet(ws, category_data: dict) -> None:
    """
    One row per (resource_type, system, code) combination.
    Sorted by resource type then count descending.
    """
    HDR = ["Resource Type", "Source Field", "Category System", "Category Code", "Display / Text", "Count"]
    ws.row_dimensions[1].height = 28
    for ci, h in enumerate(HDR, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = _font(bold=True, color=_C["hdr_fg"])
        cell.fill      = _fill(_C["hdr_bg"])
        cell.alignment = CENTER

    if not category_data:
        ws.cell(row=2, column=1, value="No category coding data found.")
        return

    # Flatten and sort
    flat_rows = []
    for rtype, counts in category_data.items():
        for (field, system, code), entry in counts.items():
            flat_rows.append((rtype, field, system, code, entry["display"], entry["count"]))
    flat_rows.sort(key=lambda x: (x[0], x[1], -x[5]))

    prev_rtype   = None
    stripe_color = _C["stripe_a"]

    for ri, (rtype, field, system, code, display, cnt) in enumerate(flat_rows, 2):
        ws.row_dimensions[ri].height = 16
        if rtype != prev_rtype:
            stripe_color = _C["stripe_a"] if stripe_color == _C["stripe_b"] else _C["stripe_b"]
            prev_rtype = rtype

        stripe = _fill(stripe_color)
        for ci, val in enumerate([rtype, field, system, code, display], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = stripe; cell.alignment = LEFT; cell.font = _font()

        cnt_cell = ws.cell(row=ri, column=6, value=cnt)
        cnt_cell.fill = stripe; cnt_cell.alignment = CENTER
        cnt_cell.font = _font(); cnt_cell.number_format = NUMFMT

    for ci, w in enumerate([28, 24, 58, 24, 50, 12], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


# ── Sheet: Data Inventory ─────────────────────────────────────────────────────

def write_data_inventory_sheet(ws, inventory_rows: list) -> None:
    """
    Exhaustive element-path frequency table across all resource types.
    """
    HDR = ["Resource Type", "Element Path", "Resource Count\nWith Element"]
    ws.row_dimensions[1].height = 36
    for ci, h in enumerate(HDR, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = _font(bold=True, color=_C["hdr_fg"])
        cell.fill      = _fill(_C["hdr_bg"])
        cell.alignment = CENTER

    if not inventory_rows:
        ws.cell(row=2, column=1, value="No data found.")
        return

    prev_rtype   = None
    stripe_color = _C["stripe_a"]

    for ri, row in enumerate(inventory_rows, 2):
        ws.row_dimensions[ri].height = 15
        if row["resource_type"] != prev_rtype:
            stripe_color = _C["stripe_a"] if stripe_color == _C["stripe_b"] else _C["stripe_b"]
            prev_rtype = row["resource_type"]

        stripe = _fill(stripe_color)
        for ci, (val, align) in enumerate([
            (row["resource_type"], LEFT),
            (row["element_path"],  LEFT),
            (row["count"],         CENTER),
        ], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = stripe
            cell.alignment = align
            cell.font      = _font()
            if ci == 3:
                cell.number_format = NUMFMT

    for ci, w in enumerate([28, 62, 20], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


# ── Dose & Route extraction ───────────────────────────────────────────────────

def _primary_med_code(resource: dict, rtype: str) -> tuple:
    """
    Return (system, code, display) of the first coding for the medication field.
    Falls back to empty strings if not present.
    """
    cc = (resource.get("code")
          if rtype == "Medication"
          else resource.get("medicationCodeableConcept"))
    if isinstance(cc, dict):
        for coding in cc.get("coding", []):
            if isinstance(coding, dict):
                return (
                    coding.get("system",  "") or "",
                    coding.get("code",    "") or "",
                    coding.get("display", "") or "",
                )
    return ("", "", "")


def _route_from_cc(route_cc) -> tuple:
    """Return (system, code, display, missing) from a route CodeableConcept."""
    if isinstance(route_cc, dict):
        for coding in route_cc.get("coding", []):
            if isinstance(coding, dict):
                return (
                    coding.get("system",  "") or "",
                    coding.get("code",    "") or "",
                    coding.get("display", "") or "",
                    False,
                )
    return ("", "", "", True)


def _dose_from_quantity(dq: dict) -> dict:
    """Unpack a SimpleQuantity / doseQuantity into dose fields."""
    return {
        "dose_type":             "Quantity",
        "dose_value":            dq.get("value"),
        "dose_unit":             dq.get("unit") or dq.get("code") or "",
        "dose_range_low_value":  None,
        "dose_range_low_unit":   None,
        "dose_range_high_value": None,
        "dose_range_high_unit":  None,
        "dose_missing":          False,
    }


def _dose_from_range(dr: dict) -> dict:
    """Unpack a doseRange into dose fields."""
    low  = dr.get("low",  {}) if isinstance(dr.get("low"),  dict) else {}
    high = dr.get("high", {}) if isinstance(dr.get("high"), dict) else {}
    return {
        "dose_type":             "Range",
        "dose_value":            None,
        "dose_unit":             None,
        "dose_range_low_value":  low.get("value"),
        "dose_range_low_unit":   low.get("unit")  or low.get("code")  or "",
        "dose_range_high_value": high.get("value"),
        "dose_range_high_unit":  high.get("unit") or high.get("code") or "",
        "dose_missing":          False,
    }


_EMPTY_DOSE = {
    "dose_type": None, "dose_value": None, "dose_unit": None,
    "dose_range_low_value": None, "dose_range_low_unit": None,
    "dose_range_high_value": None, "dose_range_high_unit": None,
    "dose_missing": True,
}


def extract_dose_route_data(resources: dict) -> tuple:
    """
    Walk MedicationRequest and MedicationAdministration resources and extract
    every dose + route entry.

    MedicationRequest  — one row per dosageInstruction element (indexed).
                         Dose comes from dosageInstruction[].doseAndRate[]
                         (doseQuantity or doseRange variants).

    MedicationAdministration — one row per resource.
                         Dose comes from dosage.dose (SimpleQuantity only).

    Returns:
        detail_rows  — all entries (present and missing)
        missing_rows — entries where route OR dose is absent
    """
    detail_rows  = []
    missing_rows = []

    # ── MedicationRequest ─────────────────────────────────────────────────────
    for r in resources.get("MedicationRequest", []):
        rid                        = r.get("id", "")
        med_sys, med_code, med_disp = _primary_med_code(r, "MedicationRequest")
        dosage_instructions        = r.get("dosageInstruction") or []

        if not dosage_instructions:
            # Resource exists but has no dosageInstruction at all
            row = {
                "resource_type": "MedicationRequest",
                "resource_id":   rid,
                "med_system":    med_sys,
                "med_code":      med_code,
                "med_display":   med_disp,
                "dosage_index":  None,
                "route_system":  "", "route_code": "", "route_display": "",
                **_EMPTY_DOSE,
                "route_missing": True,
            }
            detail_rows.append(row)
            missing_rows.append(_make_missing_row(
                "MedicationRequest", rid, med_sys, med_code, med_disp,
                missing_route=True, missing_dose=True))
            continue

        resource_missing_route = False
        resource_missing_dose  = False

        for idx, di in enumerate(dosage_instructions):
            if not isinstance(di, dict):
                continue

            # Route
            rs, rc, rd, route_missing = _route_from_cc(di.get("route"))
            if route_missing:
                resource_missing_route = True

            # Dose — walk doseAndRate, accept first doseQuantity or doseRange found
            dose_fields = dict(_EMPTY_DOSE)
            for dar in (di.get("doseAndRate") or []):
                if not isinstance(dar, dict):
                    continue
                if isinstance(dar.get("doseQuantity"), dict):
                    dose_fields = _dose_from_quantity(dar["doseQuantity"])
                    break
                if isinstance(dar.get("doseRange"), dict):
                    dose_fields = _dose_from_range(dar["doseRange"])
                    break

            if dose_fields["dose_missing"]:
                resource_missing_dose = True

            detail_rows.append({
                "resource_type": "MedicationRequest",
                "resource_id":   rid,
                "med_system":    med_sys,
                "med_code":      med_code,
                "med_display":   med_disp,
                "dosage_index":  idx + 1,
                "route_system":  rs, "route_code": rc, "route_display": rd,
                **dose_fields,
                "route_missing": route_missing,
            })

        if resource_missing_route or resource_missing_dose:
            missing_rows.append(_make_missing_row(
                "MedicationRequest", rid, med_sys, med_code, med_disp,
                missing_route=resource_missing_route,
                missing_dose=resource_missing_dose))

    # ── MedicationAdministration ──────────────────────────────────────────────
    for r in resources.get("MedicationAdministration", []):
        rid                        = r.get("id", "")
        med_sys, med_code, med_disp = _primary_med_code(r, "MedicationAdministration")
        dosage                     = r.get("dosage")

        if not isinstance(dosage, dict):
            row = {
                "resource_type": "MedicationAdministration",
                "resource_id":   rid,
                "med_system":    med_sys,
                "med_code":      med_code,
                "med_display":   med_disp,
                "dosage_index":  None,
                "route_system":  "", "route_code": "", "route_display": "",
                **_EMPTY_DOSE,
                "route_missing": True,
            }
            detail_rows.append(row)
            missing_rows.append(_make_missing_row(
                "MedicationAdministration", rid, med_sys, med_code, med_disp,
                missing_route=True, missing_dose=True))
            continue

        # Route
        rs, rc, rd, route_missing = _route_from_cc(dosage.get("route"))

        # Dose — MedAdmin uses dosage.dose (SimpleQuantity, no doseRange)
        dose_obj = dosage.get("dose")
        dose_fields = _dose_from_quantity(dose_obj) if isinstance(dose_obj, dict) else dict(_EMPTY_DOSE)

        detail_rows.append({
            "resource_type": "MedicationAdministration",
            "resource_id":   rid,
            "med_system":    med_sys,
            "med_code":      med_code,
            "med_display":   med_disp,
            "dosage_index":  1,
            "route_system":  rs, "route_code": rc, "route_display": rd,
            **dose_fields,
            "route_missing": route_missing,
        })

        if route_missing or dose_fields["dose_missing"]:
            missing_rows.append(_make_missing_row(
                "MedicationAdministration", rid, med_sys, med_code, med_disp,
                missing_route=route_missing,
                missing_dose=dose_fields["dose_missing"]))

    return detail_rows, missing_rows


def _make_missing_row(rtype, rid, med_sys, med_code, med_disp,
                      missing_route, missing_dose) -> dict:
    return {
        "resource_type": rtype,
        "resource_id":   rid,
        "med_system":    med_sys,
        "med_code":      med_code,
        "med_display":   med_disp,
        "missing_route": missing_route,
        "missing_dose":  missing_dose,
    }


# ── Sheet: Dose & Route Detail ────────────────────────────────────────────────

DR_DETAIL_HEADERS = [
    "Resource Type",       # A
    "Resource ID",         # B
    "Med Code System",     # C
    "Med Code",            # D
    "Med Display",         # E
    "Dosage\nIndex",       # F
    "Route System",        # G
    "Route Code",          # H
    "Route Display",       # I
    "Dose Type",           # J
    "Dose Value",          # K
    "Dose Unit",           # L
    "Range Low\nValue",    # M
    "Range Low\nUnit",     # N
    "Range High\nValue",   # O
    "Range High\nUnit",    # P
    "Route\nMissing",      # Q
    "Dose\nMissing",       # R
]

DR_DETAIL_WIDTHS = [24, 36, 46, 18, 46, 10, 46, 18, 36, 12, 12, 18, 14, 14, 14, 14, 10, 10]


def write_dose_route_detail_sheet(ws, detail_rows: list) -> None:
    ws.row_dimensions[1].height = 40
    for ci, h in enumerate(DR_DETAIL_HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = _font(bold=True, color=_C["hdr_fg"])
        cell.fill      = _fill(_C["hdr_bg"])
        cell.alignment = CENTER

    if not detail_rows:
        ws.cell(row=2, column=1, value="No MedicationRequest or MedicationAdministration data found.")
        return

    prev_rtype   = None
    stripe_color = _C["stripe_a"]

    for ri, row in enumerate(detail_rows, 2):
        ws.row_dimensions[ri].height = 15

        if row["resource_type"] != prev_rtype:
            stripe_color = _C["stripe_a"] if stripe_color == _C["stripe_b"] else _C["stripe_b"]
            prev_rtype = row["resource_type"]

        # Rows with any missing field get a subtle red tint on the flag columns
        base_stripe = _fill(stripe_color)

        def put(col, val, align=LEFT, fmt=None, fill=None):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.fill      = fill or base_stripe
            cell.alignment = align
            cell.font      = _font()
            if fmt:
                cell.number_format = fmt
            return cell

        put(1,  row["resource_type"])
        put(2,  row["resource_id"])
        put(3,  row["med_system"])
        put(4,  row["med_code"])
        put(5,  row["med_display"])
        put(6,  row["dosage_index"],  CENTER)
        put(7,  row["route_system"])
        put(8,  row["route_code"])
        put(9,  row["route_display"])
        put(10, row["dose_type"],     CENTER)
        put(11, row["dose_value"],    CENTER)
        put(12, row["dose_unit"],     CENTER)
        put(13, row["dose_range_low_value"],  CENTER)
        put(14, row["dose_range_low_unit"],   CENTER)
        put(15, row["dose_range_high_value"], CENTER)
        put(16, row["dose_range_high_unit"],  CENTER)

        # Flag columns — colour-coded Yes/No
        for col, flag_val in [(17, row["route_missing"]), (18, row["dose_missing"])]:
            label = "Yes" if flag_val else "No"
            fc = ws.cell(row=ri, column=col, value=label)
            fc.alignment = CENTER
            if flag_val:
                fc.fill = _fill(_C["no_bg"])
                fc.font = _font(bold=True, color=_C["no_fg"])
            else:
                fc.fill = _fill(_C["yes_bg"])
                fc.font = _font(bold=True, color=_C["yes_fg"])

    for ci, w in enumerate(DR_DETAIL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


# ── Sheet: Missing Dose or Route ──────────────────────────────────────────────

def write_missing_dose_route_sheet(ws, missing_rows: list) -> None:
    HDR = [
        "Resource Type", "Resource ID",
        "Med Code System", "Med Code", "Med Display",
        "Missing\nRoute", "Missing\nDose",
    ]
    ws.row_dimensions[1].height = 32
    for ci, h in enumerate(HDR, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = _font(bold=True, color=_C["hdr_fg"])
        cell.fill      = _fill(_C["hdr_bg"])
        cell.alignment = CENTER

    if not missing_rows:
        ws.cell(row=2, column=1,
                value="No resources with missing dose or route found — all present.")
        ws.cell(row=2, column=1).font = _font(bold=True, color=_C["yes_fg"])
        return

    # Sort: resource type → med code → resource id
    missing_rows = sorted(
        missing_rows,
        key=lambda r: (r["resource_type"], r["med_code"], r["resource_id"]),
    )

    prev_rtype   = None
    stripe_color = _C["stripe_a"]

    for ri, row in enumerate(missing_rows, 2):
        ws.row_dimensions[ri].height = 15

        if row["resource_type"] != prev_rtype:
            stripe_color = _C["stripe_a"] if stripe_color == _C["stripe_b"] else _C["stripe_b"]
            prev_rtype = row["resource_type"]

        stripe = _fill(stripe_color)

        for ci, val in enumerate([
            row["resource_type"], row["resource_id"],
            row["med_system"], row["med_code"], row["med_display"],
        ], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = stripe; cell.alignment = LEFT; cell.font = _font()

        for col, flag in [(6, row["missing_route"]), (7, row["missing_dose"])]:
            label = "Yes" if flag else "No"
            fc = ws.cell(row=ri, column=col, value=label)
            fc.alignment = CENTER
            if flag:
                fc.fill = _fill(_C["no_bg"])
                fc.font = _font(bold=True, color=_C["no_fg"])
            else:
                fc.fill = _fill(_C["yes_bg"])
                fc.font = _font(bold=True, color=_C["yes_fg"])

    # Summary counts at the bottom
    gap_row = len(missing_rows) + 3
    ws.row_dimensions[gap_row].height = 18
    n_route = sum(1 for r in missing_rows if r["missing_route"])
    n_dose  = sum(1 for r in missing_rows if r["missing_dose"])
    for ci, val in enumerate([
        "TOTALS", f"{len(missing_rows)} resources flagged",
        "", "", "",
        f"{n_route} missing route",
        f"{n_dose} missing dose",
    ], 1):
        cell = ws.cell(row=gap_row, column=ci, value=val)
        cell.font = _font(bold=True, color=_C["hdr_fg"])
        cell.fill = _fill(_C["summary_hdr"])
        cell.alignment = CENTER if ci > 1 else LEFT

    for ci, w in enumerate([24, 36, 46, 18, 46, 14, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


# ── Patient → Medication map ──────────────────────────────────────────────────
#
# Memory note: this pass works entirely on the resources dict that is already
# held in memory from the initial load.  If dataset size becomes a concern the
# natural optimisation is a second streaming pass over the ndjson files instead
# of keeping all resources in RAM — the logic below is written to make that
# refactor straightforward (no state beyond the two dicts built here).

def _ref_id(reference: str, expected_type: str = "") -> str:
    """
    Extract the logical ID from a FHIR reference string.

    Handles:
      "Patient/abc123"                     → "abc123"  (expected_type="Patient")
      "https://server/fhir/Patient/abc123" → "abc123"  (expected_type="Patient")
      "abc123"                             → "abc123"  (no type prefix, any type)

    Returns "" if the reference does not match expected_type (when supplied).
    """
    if not reference:
        return ""
    parts = reference.rstrip("/").split("/")
    if len(parts) == 1:
        # bare ID — accept regardless of expected_type
        return parts[0]
    if parts[-2] == expected_type or not expected_type:
        return parts[-1]
    return ""


def _all_med_codes(resource: dict, rtype: str,
                   med_lookup: dict | None = None) -> list:
    """
    Return every distinct (system, code) pair for a medication resource.

    Checks both the inline CodeableConcept AND any medicationReference so
    that RxNorm codes stored on a separate Medication resource are captured.
    Deduplicates on (system, code).
    """
    seen   = set()
    tuples = []

    def _add_from_cc(cc):
        if not isinstance(cc, dict):
            return
        for coding in cc.get("coding", []):
            if isinstance(coding, dict):
                system = (coding.get("system") or "").strip()
                code   = (coding.get("code")   or "").strip()
                if (system or code) and (system, code) not in seen:
                    seen.add((system, code))
                    tuples.append((system, code))

    # 1 — inline CodeableConcept
    _add_from_cc(resource.get("code") if rtype == "Medication"
                 else resource.get("medicationCodeableConcept"))

    # 2 — referenced Medication resource
    if med_lookup is not None:
        for system, code, _ in _resolve_med_ref(resource, med_lookup):
            if (system, code) not in seen:
                seen.add((system, code))
                tuples.append((system, code))

    return tuples


def build_patient_medication_map(resources: dict) -> tuple:
    """
    Build a per-patient medication code inventory from MedicationRequest and
    MedicationAdministration resources.

    Patient ID resolution order for each medication resource:
      1. subject.reference  → parsed as "Patient/{id}"
      2. subject.reference  → parsed as "Encounter/{id}", then encounter→patient map
      3. encounter.reference on the resource → encounter→patient map

    Returns:
        rows            — list of {patient_id, medication_codes} sorted by patient_id
        unresolved_count— number of medication resources whose patient could not be found
    """
    # ── Step 1: build Encounter → Patient ID lookup from Encounter resources ──
    # This is a lightweight pass — we only need subject.reference from each Encounter.
    encounter_to_patient: dict[str, str] = {}
    for enc in resources.get("Encounter", []):
        eid = enc.get("id", "")
        if not eid:
            continue
        subj_ref = ""
        subj = enc.get("subject")
        if isinstance(subj, dict):
            subj_ref = subj.get("reference", "")
        pid = _ref_id(subj_ref, "Patient")
        if pid:
            encounter_to_patient[eid] = pid

    # ── Step 2: build Medication resource lookup for reference resolution ────────
    # Resolves medicationReference pointers so RxNorm codes on separate
    # Medication resources are included alongside any inline PCC/local codes.
    med_lookup = _build_medication_lookup(resources)

    # ── Step 3: walk medication resources and collect (system, code) per patient ─
    # patient_id → set of (system, code) tuples (deduped per patient)
    patient_codes: dict[str, set] = defaultdict(set)
    unresolved = 0

    for rtype in ("MedicationRequest", "MedicationAdministration"):
        for r in resources.get(rtype, []):

            # --- resolve patient ID ---
            pid = ""

            subj = r.get("subject")
            if isinstance(subj, dict):
                ref = subj.get("reference", "")

                # Try direct Patient reference first
                pid = _ref_id(ref, "Patient")

                # If subject points to an Encounter, walk through the map
                if not pid:
                    eid = _ref_id(ref, "Encounter")
                    if eid:
                        pid = encounter_to_patient.get(eid, "")

            # Last resort: use the encounter field on the resource itself
            if not pid:
                enc_ref = r.get("encounter")
                if isinstance(enc_ref, dict):
                    eid = _ref_id(enc_ref.get("reference", ""), "Encounter")
                    if eid:
                        pid = encounter_to_patient.get(eid, "")

            if not pid:
                unresolved += 1
                continue

            # --- collect (system, code) pairs ---
            for system, code in _all_med_codes(r, rtype, med_lookup):
                patient_codes[pid].add((system, code))

    # ── Step 4: flatten to one row per (patient_id, system) ─────────────────
    # Group codes by system for each patient, then comma-delimit within each group
    from collections import defaultdict as _dd
    grouped: dict = {}
    for pid, code_pairs in patient_codes.items():
        by_system: dict = _dd(set)
        for system, code in code_pairs:
            by_system[system].add(code)
        grouped[pid] = by_system

    rows = []
    for pid in sorted(grouped):
        for system in sorted(grouped[pid]):
            codes_str = ", ".join(sorted(grouped[pid][system]))
            rows.append({
                "patient_id":  pid,
                "code_system": system,
                "codes":       codes_str,
                "code_count":  len(grouped[pid][system]),
            })

    return rows, unresolved


# ── LOINC code inventory ──────────────────────────────────────────────────────

def _find_loinc_codings(data, path: str, results: list) -> None:
    """
    Recursively walk a FHIR resource and collect every coding object whose
    system is http://loinc.org.  Stops descending once a coding is found so
    we don't re-enter its own fields.
    Arrays do not increment the path (consistent with Data Inventory behaviour).
    """
    if isinstance(data, dict):
        if data.get("system") == "http://loinc.org":
            code    = (data.get("code")    or "").strip()
            display = (data.get("display") or "").strip()
            if code:
                # Strip trailing ".coding" from path to show the CC field name
                display_path = (path[:-len(".coding")]
                                if path.endswith(".coding") else path)
                results.append((display_path, code, display))
            return   # don't recurse further into this coding object
        for key, value in data.items():
            new_path = f"{path}.{key}" if path else key
            _find_loinc_codings(value, new_path, results)
    elif isinstance(data, list):
        for item in data:
            _find_loinc_codings(item, path, results)


def build_loinc_inventory(resources: dict) -> list:
    """
    Scan all resources for LOINC codings and return one row per unique
    (resource_type, field_path, loinc_code) combination with count and display.
    """
    # {(rtype, path, code): {"display": str, "count": int}}
    inventory: dict = {}

    for rtype, rlist in resources.items():
        for r in rlist:
            hits: list = []
            _find_loinc_codings(r, "", hits)
            for path, code, display in hits:
                key = (rtype, path, code)
                if key not in inventory:
                    inventory[key] = {"display": display, "count": 0}
                inventory[key]["count"] += 1
                if not inventory[key]["display"] and display:
                    inventory[key]["display"] = display

    rows = [
        {
            "resource_type": rt,
            "field_path":    path,
            "loinc_code":    code,
            "display":       data["display"],
            "count":         data["count"],
        }
        for (rt, path, code), data in inventory.items()
    ]
    rows.sort(key=lambda r: (r["resource_type"], -r["count"], r["loinc_code"]))
    return rows


def write_loinc_sheet(ws, rows: list) -> None:
    """
    LOINC code inventory: Resource Type | Field Path | LOINC Code | Display | Count
    """
    HDR = ["Resource Type", "Field Path", "LOINC Code", "Display", "Count"]
    ws.row_dimensions[1].height = 28
    for ci, h in enumerate(HDR, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = _font(bold=True, color=_C["hdr_fg"])
        cell.fill      = _fill(_C["hdr_bg"])
        cell.alignment = CENTER

    if not rows:
        ws.cell(row=2, column=1, value="No LOINC codes found in this dataset.")
        return

    prev_rtype   = None
    stripe_color = _C["stripe_a"]

    for ri, row in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 15
        if row["resource_type"] != prev_rtype:
            stripe_color = (_C["stripe_a"] if stripe_color == _C["stripe_b"]
                            else _C["stripe_b"])
            prev_rtype = row["resource_type"]

        stripe = _fill(stripe_color)
        for ci, (val, align) in enumerate([
            (row["resource_type"], LEFT),
            (row["field_path"],    LEFT),
            (row["loinc_code"],    CENTER),
            (row["display"],       LEFT),
            (row["count"],         CENTER),
        ], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = stripe
            cell.alignment = align
            cell.font      = _font()
            if ci == 5:
                cell.number_format = NUMFMT

    # Totals footer
    footer = len(rows) + 3
    ws.row_dimensions[footer].height = 18
    n_distinct = len({r["loinc_code"] for r in rows})
    for ci, text in [
        (1, f"Resource types with LOINC codes: "
            f"{len({r['resource_type'] for r in rows})}"),
        (3, f"Distinct LOINC codes: {n_distinct:,}"),
        (5, f"Total occurrences: {sum(r['count'] for r in rows):,}"),
    ]:
        cell = ws.cell(row=footer, column=ci, value=text)
        cell.font = _font(bold=True, color=_C["hdr_fg"])
        cell.fill = _fill(_C["summary_hdr"])
        cell.alignment = LEFT

    for ci, w in enumerate([28, 46, 16, 55, 12], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


# ── Medication code co-occurrence (inferred crosswalk) ───────────────────────

def _codings_for_resource(resource: dict, rtype: str,
                           med_lookup: dict) -> list:
    """
    Return all distinct (system, code, display) tuples for a medication
    resource, combining the inline CodeableConcept and any resolved
    medicationReference so that codes from both paths are considered
    co-occurring.
    """
    seen    = set()
    results = []

    def _add(system, code, display):
        system  = (system  or "").strip()
        code    = (code    or "").strip()
        display = (display or "").strip()
        if (system or code) and (system, code) not in seen:
            seen.add((system, code))
            results.append((system, code, display))

    # 1 — inline CodeableConcept
    cc_field = "code" if rtype == "Medication" else "medicationCodeableConcept"
    cc = resource.get(cc_field)
    if isinstance(cc, dict):
        cc_text = (cc.get("text") or "").strip()
        for coding in cc.get("coding", []):
            if isinstance(coding, dict):
                _add(coding.get("system"),
                     coding.get("code"),
                     coding.get("display") or cc_text)

    # 2 — referenced Medication resource
    if rtype != "Medication":
        for sys, code, disp in _resolve_med_ref(resource, med_lookup):
            _add(sys, code, disp)

    return results


def build_code_cooccurrence(resources: dict) -> list:
    """
    Walk every Medication, MedicationRequest, and MedicationAdministration
    resource and find codes that appear together in the same CodeableConcept
    (or linked via medicationReference).

    Two codes in the same coding[] array are FHIR-equivalent representations
    of the same concept, so their pairing is an inferred cross-system mapping.

    Returns one row per unique cross-system pair, sorted by co-occurrence
    count descending.  Same-system pairs are excluded (not useful as a
    crosswalk).
    """
    med_lookup = _build_medication_lookup(resources)

    # key: (sys_a, code_a, sys_b, code_b)  — always normalised so a ≤ b
    pair_data: dict = defaultdict(lambda: {
        "count":    0,
        "disp_a":   "",
        "disp_b":   "",
        "rtypes":   set(),
    })

    for rtype in ("Medication", "MedicationRequest", "MedicationAdministration"):
        for r in resources.get(rtype, []):
            codings = _codings_for_resource(r, rtype, med_lookup)
            if len(codings) < 2:
                continue

            # Generate all cross-system pairs within this resource
            for i in range(len(codings)):
                for j in range(i + 1, len(codings)):
                    sys_a, code_a, disp_a = codings[i]
                    sys_b, code_b, disp_b = codings[j]

                    # Skip same-system pairs — not informative as a crosswalk
                    if sys_a == sys_b:
                        continue

                    # Normalise order so (A, B) and (B, A) collapse to the same key
                    if (sys_a, code_a) > (sys_b, code_b):
                        sys_a, code_a, disp_a, sys_b, code_b, disp_b = (
                            sys_b, code_b, disp_b, sys_a, code_a, disp_a
                        )

                    key  = (sys_a, code_a, sys_b, code_b)
                    data = pair_data[key]
                    data["count"] += 1
                    data["rtypes"].add(rtype)
                    # Keep the most informative display seen
                    if not data["disp_a"] and disp_a:
                        data["disp_a"] = disp_a
                    if not data["disp_b"] and disp_b:
                        data["disp_b"] = disp_b

    rows = [
        {
            "system_a":      sys_a,
            "code_a":        code_a,
            "display_a":     data["disp_a"],
            "system_b":      sys_b,
            "code_b":        code_b,
            "display_b":     data["disp_b"],
            "count":         data["count"],
            "resource_types": ", ".join(sorted(data["rtypes"])),
        }
        for (sys_a, code_a, sys_b, code_b), data in pair_data.items()
    ]

    # Sort: count desc, then system_a, code_a for stable ordering
    rows.sort(key=lambda r: (-r["count"], r["system_a"], r["code_a"]))
    return rows


# ── Sheet: Medication Code Crosswalk ─────────────────────────────────────────

def write_code_cooccurrence_sheet(ws, rows: list) -> None:
    """
    Each row is a unique cross-system code pair inferred from co-occurrence
    within the same FHIR CodeableConcept.  Higher count = stronger evidence
    that the two codes are equivalent.
    """
    HDR = [
        "System A",
        "Code A",
        "Display A",
        "System B",
        "Code B",
        "Display B",
        "Co-occurrence\nCount",
        "Confidence\nNote",
        "Found In\n(Resource Types)",
    ]
    ws.row_dimensions[1].height = 40
    for ci, h in enumerate(HDR, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = _font(bold=True, color=_C["hdr_fg"])
        cell.fill      = _fill(_C["hdr_bg"])
        cell.alignment = CENTER

    if not rows:
        ws.cell(row=2, column=1,
                value="No cross-system code pairs found — all resources use a "
                      "single coding system.")
        return

    # Determine count thresholds for confidence labelling
    counts      = [r["count"] for r in rows]
    max_count   = max(counts)
    high_thresh = max(2, max_count * 0.5)   # top 50 % of max
    med_thresh  = max(2, max_count * 0.1)   # top 10 %

    def _confidence(count):
        if count >= high_thresh:
            return "High"
        if count >= med_thresh:
            return "Medium"
        return "Low"

    for ri, row in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 15

        a_is_std = row["system_a"] in STANDARD_SYSTEMS
        b_is_std = row["system_b"] in STANDARD_SYSTEMS

        # Highlight rows where at least one side is a standard system
        if a_is_std or b_is_std:
            stripe = _fill(_C["std_bg"])
        elif ri % 2 == 0:
            stripe = _fill(_C["stripe_a"])
        else:
            stripe = _fill(_C["stripe_b"])

        def put(col, val, align=LEFT, fmt=None):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.fill      = stripe
            cell.alignment = align
            cell.font      = _font()
            if fmt:
                cell.number_format = fmt

        put(1, row["system_a"])
        put(2, row["code_a"])
        put(3, row["display_a"])
        put(4, row["system_b"])
        put(5, row["code_b"])
        put(6, row["display_b"])
        put(7, row["count"],  CENTER, NUMFMT)

        # Confidence cell — colour by level
        conf  = _confidence(row["count"])
        cc    = ws.cell(row=ri, column=8, value=conf)
        cc.alignment = CENTER
        cc.fill = stripe   # default; override below
        if conf == "High":
            cc.fill = _fill(_C["yes_bg"])
            cc.font = _font(bold=True, color=_C["yes_fg"])
        elif conf == "Medium":
            cc.fill = _fill(_C["na_bg"])
            cc.font = _font(bold=True, color=_C["na_fg"])
        else:
            cc.fill = _fill(_C["no_bg"])
            cc.font = _font(bold=True, color=_C["no_fg"])

        put(9, row["resource_types"], CENTER)

    # Summary footer
    footer_row = len(rows) + 3
    ws.row_dimensions[footer_row].height = 18
    n_std = sum(1 for r in rows
                if r["system_a"] in STANDARD_SYSTEMS
                or r["system_b"] in STANDARD_SYSTEMS)
    for ci, text in [
        (1, f"Total inferred mappings: {len(rows):,}"),
        (2, f"Pairs involving RxNorm / SNOMED / LOINC: {n_std:,}"),
        (7, f"Confidence = relative to max count ({max_count:,})"),
    ]:
        cell = ws.cell(row=footer_row, column=ci, value=text)
        cell.font = _font(bold=True, color=_C["hdr_fg"])
        cell.fill = _fill(_C["summary_hdr"])
        cell.alignment = LEFT

    col_widths = [52, 20, 46, 52, 20, 46, 14, 14, 36]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


# ── Sheet: Patient Medication Map ─────────────────────────────────────────────

def write_patient_medication_sheet(ws, rows: list, unresolved_count: int) -> None:
    """
    One row per (patient_id, code_system, code).
    Patient ID stripes change when the patient changes so it's easy to scan
    which codes belong to the same patient.
    """
    HDR = [
        "Patient ID",
        "Code System",
        "Medication Codes\n(comma-delimited)",
        "Distinct\nCode Count",
    ]
    ws.row_dimensions[1].height = 36
    for ci, h in enumerate(HDR, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = _font(bold=True, color=_C["hdr_fg"])
        cell.fill      = _fill(_C["hdr_bg"])
        cell.alignment = CENTER

    if not rows:
        ws.cell(row=2, column=1,
                value="No patient-medication data could be resolved.")
        ws.cell(row=2, column=1).font = _font(color=_C["no_fg"])
        return

    prev_pid     = None
    stripe_color = _C["stripe_a"]

    for ri, row in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 28   # taller to accommodate wrapped codes

        # Stripe alternates per patient so each patient's rows are visually grouped
        if row["patient_id"] != prev_pid:
            stripe_color = (_C["stripe_a"] if stripe_color == _C["stripe_b"]
                            else _C["stripe_b"])
            prev_pid = row["patient_id"]

        # Standard systems get the green highlight so RxNorm/SNOMED/LOINC
        # rows stand out from local/PCC rows within the same patient block
        base = (_fill(_C["std_bg"]) if row["code_system"] in STANDARD_SYSTEMS
                else _fill(stripe_color))

        for ci, (val, align) in enumerate([
            (row["patient_id"],  LEFT),
            (row["code_system"], LEFT),
            (row["codes"],       Alignment(horizontal="left", vertical="top",
                                           wrap_text=True)),
            (row["code_count"],  CENTER),
        ], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = base
            cell.alignment = align
            cell.font      = _font()
            if ci == 4:
                cell.number_format = NUMFMT

    # Summary footer
    n_patients = len({r["patient_id"] for r in rows})
    footer_row  = len(rows) + 3
    ws.row_dimensions[footer_row].height = 18
    for ci, text in [
        (1, f"Total patients: {n_patients:,}"),
        (2, f"Total (patient × system) rows: {len(rows):,}"),
        (3, f"Unresolved med resources: {unresolved_count:,}"),
    ]:
        cell = ws.cell(row=footer_row, column=ci, value=text)
        cell.font = _font(bold=True, color=_C["hdr_fg"])
        cell.fill = _fill(_C["summary_hdr"])
        cell.alignment = LEFT

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 14
    ws.freeze_panes = "A2"


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="FHIR Resource Analyzer — QI Core 6.0.0",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python fhir_analyzer.py /data/fhir_export\n"
            "  python fhir_analyzer.py /data/fhir_export -o results/analysis.xlsx\n"
            "  python fhir_analyzer.py /data/fhir_export --facility BayCare --period Aug_2025\n"
        ),
    )
    parser.add_argument("directory",
                        help="Root directory to scan recursively for .ndjson files")
    parser.add_argument("-o", "--output", default="fhir_analysis.xlsx",
                        help="Output Excel file path (default: fhir_analysis.xlsx)")
    parser.add_argument("--facility", default="",
                        help="Facility name — stamped into Summary sheet metadata")
    parser.add_argument("--period", default="",
                        help="Reporting period e.g. Aug_2025 — stamped into Summary sheet")
    parser.add_argument("--dd", default="",
                        help="Path to FHIR data dictionary Excel "
                             "(columns: Resource, Element, usedInAnalysesFlag, "
                             "conformanceVerb, bindingStrength). "
                             "Not yet implemented — provide path for future use.")
    args = parser.parse_args()

    root = Path(args.directory)
    if not root.exists() or not root.is_dir():
        print(f"ERROR: Not a valid directory: {root}")
        sys.exit(1)

    if args.dd:
        print(f"NOTE: --dd supplied ({args.dd}) but DD merge is not yet implemented. "
              f"The file will be used in a future update.")

    # ── Load ──────────────────────────────────────────────────────────────────
    print(f"Scanning:  {root.resolve()}")
    files = find_ndjson_files(root)
    print(f"Found {len(files)} .ndjson file(s)")
    if not files:
        print("No .ndjson files found. Nothing to do.")
        sys.exit(0)

    print("Loading resources…")
    resources, total_lines, parse_errors = load_resources(files)
    print(f"  {total_lines:,} lines parsed  |  {parse_errors:,} parse errors")

    relevant_types = {c["resource_type"] for c in CHECKS}
    print("\nRelevant resource counts:")
    for rtype in sorted(relevant_types):
        print(f"  {rtype:<35} {len(resources.get(rtype, [])):>8,}")
    other = sorted(set(resources) - relevant_types)
    if other:
        print(f"\nOther types found (not analysed in QI Core checks): {', '.join(other)}")

    # ── Analyse ───────────────────────────────────────────────────────────────
    print("\nCollecting profile metadata…")
    profile_counts = collect_profile_counts(resources)

    print("Running QI Core element checks…")
    results = analyze(resources)

    print("Collecting category codes…")
    category_data = collect_category_codes(resources)

    print("Building data inventory (exhaustive element scan)…")
    inventory_rows = build_data_inventory(resources)

    print("Building medication code frequency table…")
    med_rows, med_present_types = build_med_code_counts(resources)

    print("Extracting dose and route data…")
    dose_route_detail, dose_route_missing = extract_dose_route_data(resources)

    print("Building LOINC code inventory…")
    loinc_rows = build_loinc_inventory(resources)

    print("Building medication code co-occurrence crosswalk…")
    cooccurrence_rows = build_code_cooccurrence(resources)

    print("Building patient → medication map…")
    patient_med_rows, unresolved_count = build_patient_medication_map(resources)
    if unresolved_count:
        print(f"  WARNING: {unresolved_count:,} medication resource(s) could not be "
              f"linked to a patient (subject.reference missing or unresolvable).")

    # ── Expand detail rows ────────────────────────────────────────────────────
    all_detail_rows = []
    for result in results:
        rtype           = result["resource_type"]
        prof_used, prof_cnt = profile_strings(profile_counts.get(rtype, {}))
        all_detail_rows.extend(expand_result(result, prof_used, prof_cnt))

    # ── Write workbook ────────────────────────────────────────────────────────
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # 1 — Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    resource_counts_meta = {k: len(v) for k, v in resources.items()}
    for rtype, pc in profile_counts.items():
        resource_counts_meta["__profiles__" + rtype] = pc
    write_summary_sheet(ws_summary, results, resource_counts_meta,
                        args.facility, args.period)

    # 2 — FHIR Element Analysis
    ws_detail = wb.create_sheet("FHIR Element Analysis")
    write_detail_sheet(ws_detail, all_detail_rows)

    # 3 — Medication Codes
    ws_med = wb.create_sheet("Medication Codes")
    write_medication_sheet(ws_med, med_rows, med_present_types)

    # 4 — Category Codes
    ws_cat = wb.create_sheet("Category Codes")
    write_category_codes_sheet(ws_cat, category_data)

    # 5 — Data Inventory
    ws_inv = wb.create_sheet("Data Inventory")
    write_data_inventory_sheet(ws_inv, inventory_rows)

    # 6 — Dose & Route Detail
    ws_dr = wb.create_sheet("Dose & Route Detail")
    write_dose_route_detail_sheet(ws_dr, dose_route_detail)

    # 7 — Missing Dose or Route
    ws_miss = wb.create_sheet("Missing Dose or Route")
    write_missing_dose_route_sheet(ws_miss, dose_route_missing)

    # 8 — LOINC Code Inventory
    ws_loinc = wb.create_sheet("LOINC Codes")
    write_loinc_sheet(ws_loinc, loinc_rows)

    # 9 — Medication Code Crosswalk (inferred from co-occurrence)
    ws_xwalk = wb.create_sheet("Med Code Crosswalk")
    write_code_cooccurrence_sheet(ws_xwalk, cooccurrence_rows)

    # 9 — Patient Medication Map
    ws_pt = wb.create_sheet("Patient Medication Map")
    write_patient_medication_sheet(ws_pt, patient_med_rows, unresolved_count)

    wb.save(output_path)

    # ── Console summary ───────────────────────────────────────────────────────
    n_yes = sum(1 for r in results if r["present"] == "Yes")
    n_no  = sum(1 for r in results if r["present"] == "No")
    n_na  = sum(1 for r in results if r["present"] == "N/A")

    print(
        f"\nResults:     {n_yes} fully present | {n_no} have gaps | {n_na} N/A\n"
        f"Med codes:   {len(med_rows):,} distinct (system + code + display) combos\n"
        f"Cat codes:   {sum(len(v) for v in category_data.values()):,} distinct entries\n"
        f"Inventory:   {len(inventory_rows):,} distinct element paths across all types\n"
        f"Dose/Route:  {len(dose_route_detail):,} dosage entries  |  "
        f"{len(dose_route_missing):,} resources flagged missing\n"
        f"LOINC:       {len({r['loinc_code'] for r in loinc_rows}):,} distinct codes  |  "
        f"{len(loinc_rows):,} (resource type, field, code) entries\n"
        f"Crosswalk:   {len(cooccurrence_rows):,} inferred cross-system code pairs\n"
        f"Patient map: {len({r['patient_id'] for r in patient_med_rows}):,} patients  |  "
        f"{len(patient_med_rows):,} (patient × system) rows  |  "
        f"{unresolved_count:,} unresolvable\n"
        f"Report:      {output_path.resolve()}"
    )


if __name__ == "__main__":
    main()
